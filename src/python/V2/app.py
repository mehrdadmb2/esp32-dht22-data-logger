import sys
import subprocess
from datetime import datetime, timedelta

# در صورت نیاز، نصب خودکار کتابخانه‌های مورد نیاز
def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
    from PyQt6.QtWidgets import QApplication, QMainWindow, QWidget, QHBoxLayout, QVBoxLayout, QGroupBox, QLabel, QComboBox, QGridLayout, QSizePolicy
    from PyQt6.QtCore import QTimer, Qt
    from PyQt6.QtGui import QFont
except ImportError:
    install("PyQt6")
    from PyQt6.QtWidgets import QApplication, QMainWindow, QWidget, QHBoxLayout, QVBoxLayout, QGroupBox, QLabel, QComboBox, QGridLayout, QSizePolicy
    from PyQt6.QtCore import QTimer, Qt
    from PyQt6.QtGui import QFont

try:
    import requests
except ImportError:
    install("requests")
    import requests

try:
    import pyqtgraph as pg
except ImportError:
    install("pyqtgraph")
    import pyqtgraph as pg

try:
    import pandas as pd
except ImportError:
    install("pandas")
    import pandas as pd

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    install("openpyxl")
    from openpyxl import Workbook, load_workbook

try:
    import qdarkstyle
except ImportError:
    install("QDarkStyle")
    import qdarkstyle

# آدرس URL دستگاه ESP32 (بر حسب نیاز تغییر دهید)
ESP32_URL = "http://192.168.1.115/data"

DATA_FIELDS = [
    'time', 'date', 'localTemperature', 'localHumidity',
    'internetTemperature', 'internetHumidity',
    'buy_price', 'sell_price', 'gold_price',
    'ping', 'devices'
]

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("ESP32 Data Monitor")
        # تنظیم رابط کاربری اصلی
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout()
        central_widget.setLayout(main_layout)

        # پنل سمت چپ (نمودار و کنترل‌ها)
        self.setup_left_panel(main_layout)
        # پنل سمت راست (کارت‌های اطلاعات زنده)
        self.setup_right_panel(main_layout)

        # لیست برای ذخیره داده‌ها
        self.data = []
        # تایمر برای دریافت دوره‌ای داده‌ها (هر ثانیه یک‌بار)
        self.timer = QTimer()
        self.timer.timeout.connect(self.fetch_data)
        self.timer.start(1000)

    def setup_left_panel(self, parent_layout):
        left_group = QGroupBox("نمودار داده‌ها")
        left_layout = QVBoxLayout()
        left_group.setLayout(left_layout)
        parent_layout.addWidget(left_group)

        # انتخاب فیلد و بازه زمانی
        self.field_combo = QComboBox()
        self.field_combo.addItems([
            "localTemperature", "localHumidity",
            "internetTemperature", "internetHumidity",
            "buy_price", "sell_price", "gold_price",
            "ping", "devices"
        ])
        self.range_combo = QComboBox()
        self.range_combo.addItems(["1h", "1d", "1w", "1m"])
        self.field_combo.currentIndexChanged.connect(self.update_chart)
        self.range_combo.currentIndexChanged.connect(self.update_chart)

        controls_layout = QHBoxLayout()
        controls_layout.addWidget(QLabel("فیلد:"))
        controls_layout.addWidget(self.field_combo)
        controls_layout.addWidget(QLabel("بازه:"))
        controls_layout.addWidget(self.range_combo)
        left_layout.addLayout(controls_layout)

        # ویجت نمودار (پای‌کیوت‌گراف)
        self.plot_widget = pg.PlotWidget()
        self.plot_widget.setBackground('w')  # پس‌زمینه سفید (تم تیره اعمال می‌شود)
        left_layout.addWidget(self.plot_widget)

    def setup_right_panel(self, parent_layout):
        right_group = QGroupBox("اطلاعات زنده")
        right_layout = QVBoxLayout()
        right_group.setLayout(right_layout)
        parent_layout.addWidget(right_group)

        # برچسب تاریخ و زمان
        self.datetime_label = QLabel("--:--:--  ----/--/--")
        dt_font = QFont("Arial", 12, QFont.Weight.Bold)
        self.datetime_label.setFont(dt_font)
        self.datetime_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        right_layout.addWidget(self.datetime_label)

        # چینش شبکه‌ای کارت‌ها
        grid = QGridLayout()
        right_layout.addLayout(grid)

        self.field_labels = {}
        fields = [
            ('Local Temp', 'localTemperature'),
            ('Local Hum', 'localHumidity'),
            ('Internet Temp', 'internetTemperature'),
            ('Internet Hum', 'internetHumidity'),
            ('Buy Price', 'buy_price'),
            ('Sell Price', 'sell_price'),
            ('Gold Price', 'gold_price'),
            ('Ping', 'ping'),
            ('Devices', 'devices')
        ]
        for i, (name, key) in enumerate(fields):
            group = QGroupBox(name)
            vbox = QVBoxLayout()
            label = QLabel("N/A")
            label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            label.setFont(QFont("Arial", 14, QFont.Weight.Bold))
            vbox.addWidget(label)
            group.setLayout(vbox)
            row = i // 2
            col = i % 2
            grid.addWidget(group, row, col)
            self.field_labels[key] = label

    def fetch_data(self):
        try:
            response = requests.get(ESP32_URL, timeout=5)
            if response.status_code == 200:
                data = response.json()
                self.update_live_data(data)
                self.store_data_point(data)
                self.append_to_excel(data)
                self.update_chart()
        except Exception as e:
            print(f"Error fetching data: {e}")

    def update_live_data(self, data):
        # بروزرسانی برچسب تاریخ/زمان
        time_str = data.get('time', '')
        date_str = data.get('date', '')
        self.datetime_label.setText(f"{time_str}  {date_str}")
        # بروزرسانی مقادیر در کارت‌ها
        for key, label in self.field_labels.items():
            value = data.get(key, None)
            label.setText(str(value) if value is not None else "N/A")

    def store_data_point(self, data):
        # تبدیل رشته تاریخ و زمان به شیء datetime
        time_str = data.get('time', '')
        date_str = data.get('date', '')
        dt = None
        try:
            dt = datetime.strptime(date_str + ' ' + time_str, "%Y-%m-%d %H:%M:%S")
        except Exception:
            try:
                dt = datetime.strptime(date_str + ' ' + time_str, "%d-%m-%Y %H:%M:%S")
            except Exception:
                dt = datetime.now()
        point = {'datetime': dt}
        for key in DATA_FIELDS:
            if key not in ['time', 'date']:
                point[key] = data.get(key, None)
        self.data.append(point)

    def update_chart(self):
        field = self.field_combo.currentText()
        range_text = self.range_combo.currentText()
        now = datetime.now()
        if range_text == "1h":
            cutoff = now - timedelta(hours=1)
        elif range_text == "1d":
            cutoff = now - timedelta(days=1)
        elif range_text == "1w":
            cutoff = now - timedelta(weeks=1)
        else:  # 1m
            cutoff = now - timedelta(days=30)

        x = []
        y = []
        for point in self.data:
            dt = point['datetime']
            if dt >= cutoff:
                x.append(dt.timestamp())
                y.append(point.get(field, 0))
        self.plot_widget.clear()
        if x:
            self.plot_widget.plot(x, y, pen=pg.mkPen('c', width=2))
            self.plot_widget.setLabel('left', field)
            self.plot_widget.setLabel('bottom', 'Time')
            self.plot_widget.enableAutoRange('xy', True)

    def append_to_excel(self, data):
        # نام فایل اکسل بر اساس تاریخ
        date_str = data.get('date', '')
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            filename = date_obj.strftime("%Y-%m-%d") + ".xlsx"
        except Exception:
            filename = "data.xlsx"
        header = ['time', 'date', 'localTemperature', 'localHumidity',
                  'internetTemperature', 'internetHumidity',
                  'buy_price', 'sell_price', 'gold_price',
                  'ping', 'devices']
        row = [
            data.get('time', ''), data.get('date', ''),
            data.get('localTemperature', ''), data.get('localHumidity', ''),
            data.get('internetTemperature', ''), data.get('internetHumidity', ''),
            data.get('buy_price', ''), data.get('sell_price', ''),
            data.get('gold_price', ''), data.get('ping', ''), data.get('devices', '')
        ]
        try:
            wb = load_workbook(filename)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(header)  # استفاده از متد append برای ردیف‌ها:contentReference[oaicite:8]{index=8}
        ws.append(row)
        wb.save(filename)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    # اعمال تم تاریک با QDarkStyle:contentReference[oaicite:9]{index=9}:contentReference[oaicite:10]{index=10}
    app.setStyleSheet(qdarkstyle.load_stylesheet_pyqt6())
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
